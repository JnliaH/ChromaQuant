#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

This submodule contains the Theme class definition.

"""

import logging
import json
from openpyxl.styles import PatternFill, Border, Side, \
                            Alignment, Protection, Font
import os
from ..logging_and_handling import setup_logger, setup_error_logging
from typing import Literal

""" LOGGING AND HANDLING """

# Create a logger
logger = logging.getLogger(__name__)

# Format the logger
logger = setup_logger(logger)

# Get an error logging decorator
error_logging = setup_error_logging(logger)


""" CELLSTYLE CLASS"""


# Define the CellStyle class
class CellStyle:
    """
    CellStyle objects contain details about how to format each
    cell style.

    Parameters
    ----------
    font_name : str, optional
        Name of the font to use, by default 'Calibri'

    font_size : int, optional
        Size of the font, by default 11

    font_bold : bool, optional
        Whether to bold the font, by default False

    font_italic : bool, optional
        Whether to italicize the font, by default False

    font_vert_align : {'baseline', 'subscript', 'superscript', None}, optional
        Font alignment, by default None

    font_underline : {'single', 'double', 'singleAccounting', 'doubleAccounting', None}, optional
        Font underline style, by default None

    font_strike : bool, optional
        Whether to strikethrough, by default False

    font_color : str, optional
        The color of the font, by default '000000'

    fill_type : {'darkGray', 'darkUp', 'lightDown', 'darkGrid', 'darkHorizontal', 'lightTrellis', 'lightVertical', 'gray0625', 'gray125', 'lightGray', 'lightUp', 'darkDown', 'darkTrellis', 'lightGrid', 'mediumGray', 'solid', 'darkVertical', 'lightHorizontal', 'path', 'linear'}, optional
        The fill type to use for the cell background, by default None

    fill_start_color : str, optional
        The starting color for the cell background, by default 'FFFFFF'

    fill_end_color : str, optional
        The ending color for the cell background, by default 'FFFFFF'

    border_left : dict[str, str | None], optional
        Style and color of the vertical cell border. Style must be one of
        {'mediumDashed', 'mediumDashDotDot', 'dashDot', 'dashed',
        'slantDashDot', 'dashDotDot', 'thick', 'thin', 'dotted', 'double',
        'medium', 'hair', 'mediumDashDot'}. By default
        {'border_style': None, 'color': '000000'}

    border_right : dict[str, str | None], optional
        Style and color of the vertical cell border. Style must be one of
        {'mediumDashed', 'mediumDashDotDot', 'dashDot', 'dashed',
        'slantDashDot', 'dashDotDot', 'thick', 'thin', 'dotted', 'double',
        'medium', 'hair', 'mediumDashDot'}. By default
        {'border_style': None, 'color': '000000'}

    border_top : dict[str, str | None], optional
        Style and color of the vertical cell border. Style must be one of
        {'mediumDashed', 'mediumDashDotDot', 'dashDot', 'dashed',
        'slantDashDot', 'dashDotDot', 'thick', 'thin', 'dotted', 'double',
        'medium', 'hair', 'mediumDashDot'}. By default
        {'border_style': None, 'color': '000000'}

    border_bottom : dict[str, str | None], optional
        Style and color of the vertical cell border. Style must be one of
        {'mediumDashed', 'mediumDashDotDot', 'dashDot', 'dashed',
        'slantDashDot', 'dashDotDot', 'thick', 'thin', 'dotted', 'double',
        'medium', 'hair', 'mediumDashDot'}. By default
        {'border_style': None, 'color': '000000'}

    border_diagonal : dict[str, str | None], optional
        Style and color of the vertical cell border. Style must be one of
        {'mediumDashed', 'mediumDashDotDot', 'dashDot', 'dashed',
        'slantDashDot', 'dashDotDot', 'thick', 'thin', 'dotted', 'double',
        'medium', 'hair', 'mediumDashDot'}. By default
        {'border_style': None, 'color': '000000'}

    border_diagonal_up : bool, optional
        Whether to draw the diagonal from bottom-left to top-right,
        by default False

    border_diagonal_down : bool, optional
        Whether to draw the diagonal from top-left to bottom-right,
        by default False

    border_outline : bool, optional
        Whether to draw the cell outline, by default False

    border_vertical : dict[str, str | None], optional
        Style and color of the vertical cell border. Style must be one of
        {'mediumDashed', 'mediumDashDotDot', 'dashDot', 'dashed',
        'slantDashDot', 'dashDotDot', 'thick', 'thin', 'dotted', 'double',
        'medium', 'hair', 'mediumDashDot'}. By default
        {'border_style': None, 'color': '000000'}

    border_horizontal : dict[str, str | None], optional
        Style and color of the vertical cell border. Style must be one of
        {'mediumDashed', 'mediumDashDotDot', 'dashDot', 'dashed',
        'slantDashDot', 'dashDotDot', 'thick', 'thin', 'dotted', 'double',
        'medium', 'hair', 'mediumDashDot'}. By default
        {'border_style': None, 'color': '000000'}

    alignment_horizontal : {'left', 'center', 'right', 'general', 'fill', 'justify', 'centerContinuous', 'distributed'}, optional
        The horizontal alignment setting, by default 'general'

    alignment_vertical : {'top', 'center', 'bottom', 'justify', 'distributed'}, optional
        The vertical alignment setting, by default 'bottom'

    alignment_text_rotation : int, optional
        The angle, in degrees, by which to rotate text.
        Can only be integers from 0-180, by default 0

    alignment_wrap_text : bool, optional
        Whether to wrap text, by default False

    alignment_shrink_to_fit : bool, optional
        Whether to shrink the text to fit cells, by default False

    alignment_indent : float, optional
        Indent of the text in cells, by default 0

    protection_locked : bool, optional
        Whether to lock cells, by default False

    protection_hidden : bool, optional
        Whether to hide cells, by default False

    number_format : str, optional
        The formatting code to use in number formatting,
        by default 'General'

    """

    # Initialize method
    def __init__(self,
                 font_name: str = 'Calibri',
                 font_size: int = 11,
                 font_bold: bool = False,
                 font_italic: bool = False,
                 font_vert_align:
                 Literal['baseline', 'subscript', 'superscript', None] = None,
                 font_underline: str | None = None,
                 font_strike: bool = False,
                 font_color: str = '000000',
                 fill_type: str | None = None,
                 fill_start_color: str = 'FFFFFF',
                 fill_end_color: str = 'FFFFFF',
                 border_left:
                 dict[str, str | None] =
                 {'border_style': None, 'color': '000000'},
                 border_right:
                 dict[str, str | None] =
                 {'border_style': None, 'color': '000000'},
                 border_top:
                 dict[str, str | None] =
                 {'border_style': None, 'color': '000000'},
                 border_bottom:
                 dict[str, str | None] =
                 {'border_style': None, 'color': '000000'},
                 border_diagonal:
                 dict[str, str | None] =
                 {'border_style': None, 'color': '000000'},
                 border_diagonal_up: bool = False,
                 border_diagonal_down: bool = False,
                 border_outline: bool = False,
                 border_vertical:
                 dict[str, str | None] =
                 {'border_style': None, 'color': '000000'},
                 border_horizontal:
                 dict[str, str | None] =
                 {'border_style': None, 'color': '000000'},
                 alignment_horizontal: str = 'general',
                 alignment_vertical: str = 'bottom',
                 alignment_text_rotation: int = 0,
                 alignment_wrap_text: bool = False,
                 alignment_shrink_to_fit: bool = False,
                 alignment_indent: float = 0,
                 protection_locked: bool = False,
                 protection_hidden: bool = False,
                 number_format: str = 'General'):

        # Define dictionaries for each style property
        # Set Font values
        self._font = {'name': font_name,
                      'size': font_size,
                      'bold': font_bold,
                      'italic': font_italic,
                      'vertAlign': font_vert_align,
                      'underline': font_underline,
                      'strike': font_strike,
                      'color': font_color
                      }

        # Set PatternFill values
        self._fill = {'fill_type': fill_type,
                      'start_color': fill_start_color,
                      'end_color': fill_end_color
                      }

        # Set Border values
        self._border = {'left':
                        border_left,
                        'right':
                        border_right,
                        'top':
                        border_top,
                        'bottom':
                        border_bottom,
                        'diagonal':
                        border_diagonal,
                        'diagonalUp':
                        border_diagonal_up,
                        'diagonalDown':
                        border_diagonal_down,
                        'outline':
                        border_outline,
                        'vertical':
                        border_vertical,
                        'horizontal':
                        border_horizontal,
                        }

        # Set Alignment values
        self._alignment = {'horizontal': alignment_horizontal,
                           'vertical': alignment_vertical,
                           'text_rotation': alignment_text_rotation,
                           'wrap_text': alignment_wrap_text,
                           'shrink_to_fit': alignment_shrink_to_fit,
                           'indent': alignment_indent
                           }

        # Set Protection values
        self._protection = {'locked': protection_locked,
                            'hidden': protection_hidden
                            }

        # Set Number format
        self._number_format = number_format

        # Create empty column pointer
        self.column_pointer = ''

        # Get a list of acceptable border sides
        self.border_side_list = ['left', 'right', 'top', 'bottom', 'diagonal',
                                 'vertical', 'horizontal']

    # Method to add a pointer to the CellStyle
    @error_logging
    def point_to(self,
                 column_id: str):

        # Set the column pointer
        self.column_pointer = column_id

        return None

    """ PROPERTIES """
    # Font
    # Getter
    @property
    def font(self):
        # Return the Font object as evaluated at get
        return Font(**self._font)

    # Setter
    @font.setter
    def font(self, value):
        self._font = value

    # Fill
    # Getter
    @property
    def fill(self):
        # Return the Fill object as evaluated at get
        return PatternFill(**self._fill)

    # Setter
    @fill.setter
    def fill(self, value):
        self._fill = value

    # Border
    # Getter
    @property
    def border(self):
        # Get a new dictionary for each side with values as Side
        # with arguments as border dictionary values, checking that
        # each argument processed is a valid side
        border_sides = {border_side: Side(**border_side_dict)
                        for border_side, border_side_dict
                        in self._border.items()
                        if border_side in self.border_side_list}
        # Get a dictionary of non-side arguments
        border_nonsides = {border_key: border_val
                           for border_key, border_val
                           in self._border.items()
                           if border_key not in self.border_side_list}
        # Get a merged kwargs dictionary
        border_kwargs = border_sides | border_nonsides
        # Return the Border object as evaluated at get
        return Border(**border_kwargs)

    # Setter
    @border.setter
    def border(self, value):
        self._border = value

    # Alignment
    # Getter
    @property
    def alignment(self):
        # Return the Alignment object as evaluated at get
        return Alignment(**self._alignment)

    # Setter
    @alignment.setter
    def alignment(self, value):
        self._alignment = value

    # Protection
    # Getter
    @property
    def protection(self):
        # Return the Protection object as evaluated at get
        return Protection(**self._protection)

    # Setter
    @protection.setter
    def protection(self, value):
        self._protection = value

    # Number formatt
    # Getter
    @property
    def number_format(self):
        return self._number_format

    # Setter
    @number_format.setter
    def number_format(self, value):
        self._number_format = value


""" CHARTSTYLE CLASS """


# Define the ChartStyle class
class ChartStyle:
    """
    ChartStyle objects contain details about how to format charts.

    Parameters
    ----------
    chart_outline : bool, optional
        Whether to draw an outline around the chart, by default True

    chart_vary_colors : bool, optional
        Whether to vary colors in individual series, by default True

    chart_rounded_corners : bool, optional
        Whether to round outer chart corners, by default True

    plot_area_draw_outline : bool, optional
        Whether to draw an outline around the plot area, by default False

    plot_area_outline_width : float | None, optional
        The width of the plot area outline in points, by default None

    plot_area_outline_style : {'thinThick', 'tri', 'sng', 'thickThin', 'dbl', None}, optional
        The plot area outline style, by default None

    plot_area_outline_color : str | None, optional
        The plot area outline color, by default None

    legend_position : {'r', 'l', 't', 'b', 'tr', None}, optional
        The legend position, by default None

    legend_overlay : bool, optional
        Whether to overlay the legend, by default True

    title : str | None, optional
        The chart title, by default None

    title_overlay : bool, optional
        Whether to overlay the chart title, by default True

    title_font_name : str | None, optional
        The latin font to use for the chart title, by default None

    title_font_size : float | int | None, optional
        The chart title font size, by default None

    x_title : str | None, optional
        The x-axis title, by default None

    x_overlay : bool, optional
        Whether to overlay the x-axis title, by default True

    x_font_name : str | None, optional
        The latin font to use for the x-axis title, by default None

    x_font_size : float | int | None, optional
        The x-axis title font size, by default None

    x_draw_major_grid : bool, optional
        Whether to draw major x-axis gridlines, by default True

    x_draw_minor_grid : bool, optional
        Whether to draw minor x-axis gridlines, by default False

    x_show_numbers : bool, optional
        Whether to show x-axis numbers, by default False

    x_tick_major_style : {'in', 'out', 'cross', None}, optional
        The x-axis major tick mark style, by default None

    x_tick_minor_style : {'in', 'out', 'cross', None}, optional
        The x-axis minor tick mark style, by default None

    x_tick_major_unit : float | None, optional
        The x-axis major unit, by default None

    x_tick_minor_unit : float | None, optional
        The x-axis minor unit, by default None

    y_title : str | None, optional
        The y-axis title, by default None

    y_overlay : bool, optional
        Whether to overlay the y-axis title, by default True

    y_font_name : str | None, optional
        The latin font to use for the y-axis title, by default None

    y_font_size : float | int | None, optional
        The y-axis title font size, by default None

    y_draw_major_grid : bool, optional
        Whether to draw major y-axis gridlines, by default True

    y_draw_minor_grid : bool, optional
        Whether to draw minor y-axis gridlines, by default False

    y_show_numbers : bool, optional
        Whether to show y-axis numbers, by default False

    y_tick_major_style : {'in', 'out', 'cross', None}, optional
        The y-axis major tick mark style, by default None

    y_tick_minor_style : {'in', 'out', 'cross', None}, optional
        The y-axis minor tick mark style, by default None

    y_tick_major_unit : float | None, optional
        The y-axis major unit, by default None

    y_tick_minor_unit : float | None, optional
        The y-axis minor unit, by default None

    """

    # Initialize method
    def __init__(self,
                 chart_outline: bool = True,
                 chart_vary_colors: bool = True,
                 chart_rounded_corners: bool = True,
                 plot_area_draw_outline: bool = False,
                 plot_area_outline_width: float | None = None,
                 plot_area_outline_style: str | None = None,
                 plot_area_outline_color: str | None = None,
                 legend_position: str | None = None,
                 legend_overlay: bool = True,
                 title: str | None = None,
                 title_overlay: bool = True,
                 title_font_name: str | None = None,
                 title_font_size: float | int | None = None,
                 x_title: str | None = None,
                 x_overlay: bool = True,
                 x_font_name: str | None = None,
                 x_font_size: float | int | None = None,
                 x_draw_major_grid: bool = True,
                 x_draw_minor_grid: bool = False,
                 x_show_numbers: bool = False,
                 x_tick_major_style: str | None = None,
                 x_tick_minor_style: str | None = None,
                 x_tick_major_unit: float | None = None,
                 x_tick_minor_unit: float | None = None,
                 y_title: str | None = None,
                 y_overlay: bool = True,
                 y_font_name: str | None = None,
                 y_font_size: float | int | None = None,
                 y_draw_major_grid: bool = True,
                 y_draw_minor_grid: bool = False,
                 y_show_numbers: bool = False,
                 y_tick_major_style: str | None = None,
                 y_tick_minor_style: str | None = None,
                 y_tick_major_unit: float | None = None,
                 y_tick_minor_unit: float | None = None
                 ):

        # Set chart format
        self._chart = \
            {
                'chart_outline': chart_outline,
                'vary_colors': chart_vary_colors,
                'rounded_corners': chart_rounded_corners
            }

        # Set plot area format
        self._plot_area = \
            {
                'draw_outline': plot_area_draw_outline,
                'outline_width': plot_area_outline_width,
                'outline_style': plot_area_outline_style,
                'outline_color': plot_area_outline_color
            }

        # Set legend format
        self._legend = \
            {
                'position': legend_position,
                'overlay': legend_overlay
            }

        # Set title format
        self._title = \
            {
                'title': title,
                'overlay': title_overlay,
                'font_name': title_font_name,
                'font_size': title_font_size
            }

        # Set x-axis format
        self._x_axis = \
            {
                'title': x_title,
                'overlay': x_overlay,
                'font_name': x_font_name,
                'font_size': x_font_size,
                'draw_major_grid': x_draw_major_grid,
                'draw_minor_grid': x_draw_minor_grid,
                'show_numbers': x_show_numbers,
                'tick_major_style': x_tick_major_style,
                'tick_minor_style': x_tick_minor_style,
                'tick_major_unit': x_tick_major_unit,
                'tick_minor_unit': x_tick_minor_unit
            }

        # Set y-axis format
        self._y_axis = \
            {
                'title': y_title,
                'overlay': y_overlay,
                'font_name': y_font_name,
                'font_size': y_font_size,
                'draw_major_grid': y_draw_major_grid,
                'draw_minor_grid': y_draw_minor_grid,
                'show_numbers': y_show_numbers,
                'tick_major_style': y_tick_major_style,
                'tick_minor_style': y_tick_minor_style,
                'tick_major_unit': y_tick_major_unit,
                'tick_minor_unit': y_tick_minor_unit
            }

        """ PROPERTIES """
        # Chart
        # Getter
        @property
        def chart(self):
            return self._chart

        # Setter
        @chart.setter
        def chart(self, value):
            self._chart = value

        # Plot area
        # Getter
        @property
        def plot_area(self):
            return self._plot_area

        # Setter
        @plot_area.setter
        def plot_area(self, value):
            self._plot_area = value

        # Legend
        # Getter
        @property
        def legend(self):
            return self._legend

        # Setter
        @legend.setter
        def legend(self, value):
            self._legend = value

        # Title
        # Getter
        @property
        def title(self):
            return self._title

        # Setter
        @title.setter
        def title(self, value):
            self._title = value

        # X-axis
        # Getter
        @property
        def x_axis(self):
            return self._x_axis

        # Setter
        @x_axis.setter
        def x_axis(self, value):
            self._x_axis = value

        # Y-axis
        # Getter
        @property
        def y_axis(self):
            return self._y_axis

        # Setter
        @y_axis.setter
        def y_axis(self, value):
            self._y_axis = value


""" THEME CLASS """


# Define the Theme class
class Theme:
    """
    Theme objects contain details about how to format reports from
    exporting Results.
    """

    # Initialize method
    def __init__(self):

        # Define a list of cell styles
        self.cell_styles = ['header', 'subheader', 'body']

        # Define a list of chart styles
        self.chart_styles = \
            ['areaChart', 'area3DChart', 'barChart', 'bar3DChart',
             'bubbleChart', 'lineChart', 'line3DChart', 'pieChart',
             'pie3DChart', 'doughnutChart', 'custSplit', 'ofPieChart',
             'radarChart', 'scatterChart', 'stockChart', 'surface3DChart']

        # For every cell style...
        for style in self.cell_styles:
            # Create a style group
            setattr(self, style, CellStyle())

        # For every chart style...
        for style in self.chart_styles:
            # Create a style group
            setattr(self, style, ChartStyle())

        # Create a dictionary for custom column style groups
        self.columns = {}

        # Get the current script path
        script_dir = os.path.dirname(os.path.abspath(__file__))

        # Load the default cqtheme
        self.import_theme(
            os.path.join(script_dir, 'themes', 'default.cqtheme')
            )

    # Method to add a new style group for a column
    # NOTE: only used for Tables and Breakdowns
    @error_logging
    def add_column_style_group(self,
                               group_name: str,
                               column_id: str = '',
                               style_group: CellStyle = CellStyle(),
                               ):

        # Add the style_group as an attribute to the columns attribute
        self.columns[group_name] = style_group

        # If the column value is not empty
        if column_id:
            # Add a pointer to the style group
            self.columns[group_name].point_to(column_id)

        # Otherwise, pass
        else:
            pass

        return None

    # Method to import theme from .cqtheme
    @error_logging
    def import_theme(self, path):

        # Load the .cqtheme at path
        with open(path, 'r') as f:
            theme_dict = json.load(f)

        # For every group in the theme...
        for group, group_dict in theme_dict.items():

            # If the group is a cell or chart group...
            if group in self.cell_styles or group in self.chart_styles:

                # Add each style property to the style group
                for style_property, property_val in group_dict.items():
                    setattr(getattr(self, group), style_property, property_val)

            # Otherwise, pass
            else:
                pass

        return None
