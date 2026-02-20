#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

The Formula Tools module contains a number of functions used internally
by the Formula class.

"""

import re
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from typing import Any

""" FUNCTIONS """


# Function for checking that passed formula starts with '='
def check_formula_starts_with_equals(formula: str) -> bool:
    """
    Function that checks whether a formula string starts with an equals sign.

    Parameters
    ----------
    formula : str
        Excel formula string.

    Returns
    -------
    tf: bool
        True if formula starts with equals sign, False if not.
    """

    # Initialize tf
    tf = False

    # If a match exists, set tf to True
    if not re.search('^=', formula) is None:
        tf = True
    # Otherwise, pass
    else:
        pass

    return tf


# Function to get a column letter based on table
# starting column and column's index within a table
def get_column_letter_from_table(table_reference: dict[str, Any],
                                 column_name: str,
                                 start_cell_column: int) -> str:
    """
    Function that gets a column letter for a column in a Table.

    Parameters
    ----------
    table_reference : dict
        A Table's reference attribute.
    column_name : str
        The name of the column of interest.
    start_cell_column : int
        The index of the Table's starting cell.

    Returns
    -------
    column_letter: str
        The reference letter for the column.

    """

    # Find the column's index within the DataFrame
    column_index = \
        table_reference['columns'].index(column_name)

    # Add the column's index to the starting column
    # and get the column letter equivalent
    column_letter = \
        get_column_letter(start_cell_column + column_index)

    return column_letter


# Function to construct value reference string
def get_value_reference_string(value_reference: dict[str, Any]) -> str:
    """
    Function that constructs a value's cell reference from its reference
    attribute.

    Parameters
    ----------
    value_reference : dict[str, Any]
        A Value's reference attribute.

    Returns
    -------
    reference: str
        A constructed cell reference.

    """

    # Get value's sheet string
    value_sheet = \
        f"'{value_reference['sheet']}'!"
    # Get value's cell string
    value_cell = \
        f"{value_reference['cell']}"

    # Get the reference string
    reference = value_sheet + value_cell

    return reference


# Function to get the starting row and column of a table
def get_table_start_coords(table_reference: dict[str, Any]) -> tuple[int, int]:
    """
    Function that returns the starting row and column of a table.

    Parameters
    ----------
    table_reference : dict[str, Any]
        A Table's reference attribute.

    Returns
    -------
    start_row : int
        The starting row index.
    start_column : int
        The starting column index.

    """

    # Get the table's starting cell
    table_start_cell = table_reference['start_cell']

    # Get the starting cell's row and column
    start_row, start_column = \
        coordinate_to_tuple(table_start_cell)

    return start_row, start_column


# Function to replace formula raw insert with reference
def replace_insert(formula: str, raw: str, reference: str) -> str:
    """
    Function that returns a formula after replacing some passed substring
    with another substring.

    Parameters
    ----------
    formula : str
        Formula containing a substring to be replaced.
    raw : str
        Substring to be replaced.
    reference : str
        Substring to substitute for raw.

    Returns
    -------
    new_formula : str
        Formula after substring replacement.
    """

    # Replace pointer substring with value's reference
    new_formula = \
        formula.replace(raw, reference)

    return new_formula


# Function to get a range from passed table pointers
def table_column_to_range(table_reference: dict[str, Any],
                          column_name: str) -> str:
    """
    Function that gets a Table column's range string.

    Parameters
    ----------
    table_reference : dict[str, Any]
        A Table's reference attribute.
    column_name : str
        The name of a column of interest in Table.

    Returns
    -------
    column_range : str
        Range for Table column.

    """

    # Get the starting coordinates
    start_row, start_column = get_table_start_coords(table_reference)

    # Get letter for output column
    column_letter = \
        get_column_letter_from_table(table_reference,
                                     column_name,
                                     start_column)

    # Get the final row of the table
    end_row = start_row + table_reference['length']

    # Construct the range string
    column_range = (
        f'${column_letter}${start_row + 1}:'
        f'${column_letter}${end_row + 1}')

    return column_range
