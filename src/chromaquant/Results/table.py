#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
COPYRIGHT STATEMENT:

ChromaQuant – A quantification software for complex gas chromatographic data

Copyright (c) 2026, by Julia Hancock
              Affiliation: Dr. Julie Elaine Rorrer
              URL: https://www.rorrerlab.com/

License: BSD 3-Clause License

---

CLASS DEFINITION FOR TABLES

Julia Hancock
Started 01-12-2025

"""

import pandas as pd
from openpyxl.utils.cell import get_column_letter, \
                                column_index_from_string, \
                                coordinate_from_string
from .dataset import DataSet
from .. import logging_and_handling as lah


""" LOGGING AND HANDLING """

# Get the logger
logger = lah.setup_logger()

# Get an error logging decorator
error_logging = lah.setup_error_logging(logger)

""" CLASS """


# Define the Table class
class Table(DataSet):

    def __init__(self,
                 data_frame=pd.DataFrame(),
                 *args,
                 **kwargs):

        # Run DataSet initialization
        super().__init__(*args, **kwargs)

        # Initialize data attribute
        self.data = data_frame

        # Update the table
        self.update_table()

    """ PROPERTIES """
    # Define the references property
    @property
    def references(self):
        return self._references

    """ METHODS """
    # Method to update column references
    @error_logging
    def update_references(self):

        # Get the indices for the start cell
        start_col_index, start_row_index = \
            self.get_cell_indices(self._start_cell)

        # For every column in columns...
        for column in self.columns:

            # Get the current column's letter
            col_letter = \
                self.get_column_letter_wrt_start_cell(column,
                                                      self.data,
                                                      start_col_index)

            # Update the references object
            self._references[column] = \
                {'column_letter': col_letter,
                 'start_row': start_row_index + 1,
                 'end_row': start_row_index + self.length,
                 'sheet': self._sheet,
                 'range': (f"'{self._sheet}'!"
                           f"${col_letter}${start_row_index + 1}:"
                           f"${col_letter}${start_row_index + self.length}")}

        return None

    # Method for updating otherwise static Table attributes
    @error_logging
    def update_table(self):

        # Update the length header
        self.length = \
            len(self.data)

        # Update the column header
        self.columns = \
            self.data.columns.tolist()

        # Initialize the references object
        self._references = {}

        # Try to update the references
        # NOTE: will not work if there is no valid sheet or start_cell
        try:
            self.update_references()
        except AttributeError as e:
            logger.info(e)
            pass

        return None

    # Method to add a new column to a table
    @error_logging
    def add_table_column(self,
                         column_name,
                         column_values=float('nan')):

        # Set every entry in the column to column_values
        # If column_values is an iterable, this will iterate
        # through it, otherwise it will set every entry to
        # the same value
        self.data[column_name] = column_values

        # Update the table
        self.update_table()

        return None

    """ STATIC METHODS """
    # Function to return the row and column indices of a given cell
    @staticmethod
    def get_cell_indices(cell):

        # Split the cell string into individual coordinates
        col_letter, row_index = \
            coordinate_from_string(cell)

        # Get the starting column's index from its letter
        col_index = column_index_from_string(col_letter)

        return col_index, row_index

    # Function to get the letter coordinate of a column in a DataFrame
    # with respect to a starting column index
    @staticmethod
    def get_column_letter_wrt_start_cell(column, df, start_col_index):

        # Get the current column's letter
        col_letter = \
            get_column_letter(start_col_index +
                              df.columns.get_loc(column))

        return col_letter
