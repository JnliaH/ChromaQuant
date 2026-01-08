#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
COPYRIGHT STATEMENT:

ChromaQuant – A quantification software for complex gas chromatographic data

Copyright (c) 2026, by Julia Hancock
              Affiliation: Dr. Julie Elaine Rorrer
              URL: https://www.rorrerlab.com/

License: BSD 3-Clause License

---

CLASS DEFINITION FOR FORMULA

Julia Hancock
Started 01-07-2026

"""

import re
import pandas as pd
from openpyxl.utils import get_column_letter, coordinate_to_tuple
from chromaquant import import_local_packages as ilp
from chromaquant import logging_and_handling as lah

""" LOGGING AND HANDLING """

# Get the logger
logger = lah.setup_logger()

# Get an error logging decorator
error_logging = lah.setup_error_logging(logger)

""" LOCAL PACKAGES """

# Get absolute directories for subpackages
subpack_dir = ilp.get_local_package_directories()

# Import all local packages
hd = ilp.import_from_path("hd", subpack_dir['Handle'])

""" CLASSES """


# Define the formula class
class Formula:

    # Initialize class
    def __init__(self, formula, *args, **kwargs):
        """
        __init__ _summary_

        Parameters
        ----------
        formula : str
            Passed formula with or without inserts and pointers

        """

        # Extract the formula
        self.formula = formula

        # Get the formula's cell inserts
        self.get_formula_cell_inserts()

        # Extract passed keywork arguments
        for key, value in kwargs.items():
            # Define a prefix
            prefix = 'signalkwarg'
            # Create an attribute name using the defined prefix
            attribute_name = prefix + '_' + key
            # Set attribute using name and key-value pair
            setattr(self, attribute_name, value)
    
    # Function for checking that passed formula starts with '='
    @staticmethod
    def check_formula_starts_with_equals(formula):

        # Initialize tf
        tf = False

        # If a match exists, set tf to True
        if not re.search('^=', formula) is None:
            tf = True
        # Otherwise, pass
        else:
            pass

        return tf

    # Function to get cases in a formula where cells need to be inserted
    @error_logging
    def get_formula_cell_inserts(self):

        # Initialize a list to contain parsed pipe contents
        insert_list = []

        # If the formula contains at least one pipe character...
        if '|' in self.formula:

            # Get the number of pipes in the formula
            pipe_count = self.formula.count('|')

            # If the number of pipes is divisible by two...
            if pipe_count % 2 == 0:

                # Get a list of every substring within pipes
                insert_list = [{'start': match.start(),
                                'end': match.end(),
                                'raw': match.group(),
                                'pipe': [],
                                'pointers': {}}
                               for match in re.finditer(r'\|(.*?)\|', self.formula)]

                # For each pipe substring...
                for i in range(len(insert_list)):

                    # Split each substring if it contains commas
                    insert_list[i]['pipe'] = \
                        insert_list[i]['raw'].split(',')

                    # For each newly-split substring...
                    for j in range(len(insert_list[i]['pipe'])):

                        # If the substring contains one colon...
                        if insert_list[i]['pipe'][j].count(':') == 1:

                            # Split the substring by that colon
                            key_value_list = \
                                insert_list[i]['pipe'][j].split(':')

                            # Strip the key and value of pipes and whitespace
                            key_value_list = \
                                [k.strip('|').strip() for k in key_value_list]

                            # Add the key-value pair to the pointers dictionary
                            insert_list[i]['pointers'][key_value_list[0]] = \
                                key_value_list[1]

                        # Otherwise, raise error
                        else:
                            raise ValueError('At least one pointer contains an'
                                             'unexpected number of colons')

                    # Remove the pipe key-value pairs
                    del insert_list[i]['pipe']

            # Otherwise, raise an error
            else:
                raise ValueError('Formula contains at least'
                                 'one hanging pipe delimeter:'
                                 f' "{self.formula}"')

        # Otherwise, pass
        else:
            pass

        # Save the insert list to a class attribute
        self.insert_list = insert_list

        return None

    # Function to construct value reference string
    @staticmethod
    def get_value_reference_string(value_header):

        # Get value's sheet string
        value_sheet = \
            f"'{value_header['sheet']}'!"
        # Get value's cell string
        value_cell = \
            f"{value_header['cell']}"

        # Get the reference string
        reference = value_sheet + value_cell

        return reference

    # Function to replace formula raw insert with reference
    @staticmethod
    def replace_insert(formula, raw, reference):

        # Replace pointer substring with value's reference
        new_formula = \
            formula.replace(raw, reference)

        return new_formula

    # Function to get the starting row and column of a table
    @staticmethod
    def get_table_start_coords(table_header):

        # Get the table's starting cell
        table_start_cell = table_header['start_cell']

        # Get the starting cell's row and column
        start_row, start_column = \
            coordinate_to_tuple(table_start_cell)

        return start_row, start_column

    # Function to get a column letter based on table
    # starting column and column's index within a table
    @staticmethod
    def get_column_letter_from_table(
            table_headers, column_name, start_cell_column):

        # Find the column's index within the DataFrame
        column_index = \
            table_headers['columns'].index(column_name)

        # Add the column's index to the starting column
        # and get the column letter equivalent
        column_letter = \
            get_column_letter(start_cell_column + column_index)

        return column_letter

    # Function to get a range from passed table pointers
    def table_column_to_range(self, table, column_key):

        # Get the starting coordinates
        start_row, start_column = self.get_table_start_coords(table)

        # Get letter for output column
        column_letter = \
            self.get_column_letter_from_table(table, column_key, start_column)

        # Get the final row of the table
        end_row = start_row + table['length']

        # Construct the range string
        column_range = (
            f'${column_letter}${start_row + 1}:'
            f'${column_letter}${end_row + 1}')

        return column_range

    # Function to process formula inserts
    @error_logging
    def process_table_formula_inserts(self, value_headers, table_headers, output_pointers):
        """
        process_formula_inserts

        Parameters
        ----------
        value_headers : dict
            Dictionary of headers by value name.
        table_headers : dict
            Dictionary of headers by table name.
        output_pointers : _type_
            Pointer(s) to where new formulas will go.

        Returns
        -------
        new_formula : list or str
            New formula or list of new formulas.

        Raises
        ------
        ValueError
            'Insert list contains non-key or non-table elements'
        ValueError
            'Formula missing necessary keys'
        ValueError
            'Passed output pointers do not point to either a value or table'
        """

        # If the insert list is not None (i.e., there are inserts)
        if self.insert_list is not None:

            # If function was passed a table to output to
            if 'table' in output_pointers \
                 and 'key' in output_pointers:

                # Try to get the output table length:
                try:
                    output_table_length = \
                        table_headers[output_pointers['table']]['length']

                # If can't get the length, set to five
                except KeyError:
                    output_table_length = 5

                # If there is at least one table among inserts...
                if any('table' in insert.get('pointers', {})
                       and 'key' in insert.get('pointers', {})
                       for insert in self.insert_list):

                    # Get the length of the longest named table
                    max_table_length = max(
                        [table_headers[insert['pointers']['table']]['length']
                            for insert in self.insert_list
                            if 'table' in insert.get('pointers', {})]
                    )

                    # If the max_table_length is greater than 1, set
                    # the output_table_length to the max_table_length
                    if max_table_length > 1:
                        output_table_length = max_table_length

                    # Otherwise, pass
                    else:
                        pass

                    # Initialize a new formula list
                    new_formula = [self.formula for i in range(output_table_length)]

                    # For every insert...
                    for insert in self.insert_list:

                        # If the insert has a table and key pointer...
                        if 'table' in insert['pointers'] \
                           and 'key' in insert['pointers']:

                            # Get the insert's table and key pointers
                            table_name = insert['pointers']['table']
                            column_name = insert['pointers']['key']

                            # Get the current table header
                            insert_table = table_headers[table_name]

                            # If the insert has range pointer equal to true...
                            if 'range' in insert['pointers'] and \
                               insert['pointers']['range'].capitalize() \
                               == 'True':

                                # Get a range substring for each
                                # new entry in the output table
                                insert['reference'] = \
                                    self.table_column_to_range(insert_table,
                                                               column_name)

                                # Replace the insert's raw substring
                                # for every formula in the new formula list
                                new_formula = \
                                    [self.replace_insert(one_formula,
                                                    insert['raw'],
                                                    insert['reference'])
                                     for one_formula in new_formula]

                            # Otherwise...
                            else:

                                # Get the table's starting coordinates
                                start_row, start_column = \
                                    self.get_table_start_coords(insert_table)

                                # Get letter for output column
                                column_letter = \
                                    self.get_column_letter_from_table(
                                            insert_table,
                                            column_name,
                                            start_column)

                                # Get a list of insert references for each
                                # row iterated over
                                insert['reference'] = \
                                    [f'{column_letter}{row}' for row in
                                     range(start_row + 1,
                                           start_row + output_table_length + 1)
                                     ]

                                # Replace the insert's raw substring
                                # for every formula in the new formula list
                                new_formula = \
                                    [self.replace_insert(new_formula[i],
                                                    insert['raw'],
                                                    insert['reference'][i])
                                     for i in range(output_table_length)]

                        # Otherwise, if the insert has a key pointer...
                        elif 'key' in insert['pointers']:

                            # Get the value's reference string
                            insert['reference'] = \
                                self.get_value_reference_string(
                                    value_headers[insert['pointers']['key']]
                                    )

                            # Replace the insert's raw substring
                            # for every formula in the new formula list
                            new_formula = \
                                [self.replace_insert(one_formula,
                                                     insert['raw'],
                                                     insert['reference'])
                                    for one_formula in new_formula]

                        # Otherwise, raise an error
                        else:
                            raise ValueError('Insert list contains '
                                             'non-key or non-table elements')

                # Otherwise, if there are only values among inserts...
                elif all('key' in insert.get('pointers', {})
                         for insert in self.insert_list):

                    # Initialize a new formula list
                    new_formula = [self.formula for i in range(output_table_length)]

                    # Initialize a string formula to use in iterations
                    iterate_formula = self.formula

                    # For every insert...
                    for insert in self.insert_list:

                        # Get the value's reference string
                        insert['reference'] = \
                            self.get_value_reference_string(
                                insert['pointers']['key']
                            )

                        # Replace insert substring with value's reference
                        iterate_formula = \
                            self.replace_insert(iterate_formula,
                                                insert['raw'],
                                                insert['reference'])

                    # Set the new formula list to be iterate_formula for
                    # every entry in the output_table's column
                    new_formula = [iterate_formula
                                   for row in range(output_table_length)]

                # Otherwise, raise an exception
                else:
                    raise ValueError(
                        'Insert list contains non-key or non-table elements')



            # Otherwise, raise an error
            else:
                raise ValueError('Passed output pointers do not'
                                 ' point to either a value or table')

        # Otherwise, pass
        else:
            pass

        # Set the new formula attribute to the result
        self.new_formula = new_formula

        return None
    
    def process_value_formula_inserts(self, value_headers, table_headers):

        # Initialize a new formula
        new_formula = self.formula

        # For every insert...
        for insert in self.insert_list:

            # If the insert has a table and key pointer...
            if 'table' in insert['pointers'] \
                and 'key' in insert['pointers']:

                # Get the insert's table and key pointers
                table_name = insert['pointers']['table']
                column_name = insert['pointers']['key']

                # Get the table header
                insert_table = table_headers[table_name]

                # Get a range substring and save to insert['reference']
                insert['reference'] = \
                    self.table_column_to_range(insert_table, column_name)

                # Replace pointer substring with range substring
                new_formula = \
                    new_formula.replace(
                        insert['raw'],
                        insert['reference'])

            # Otherwise, if insert has a key pointer...
            elif 'key' in insert['pointers']:

                # Get the value's reference string
                insert['reference'] = \
                    self.get_value_reference_string(
                        value_headers[insert['pointers']['key']]
                        )

                # Replace insert substring with value's reference
                new_formula = \
                    self.replace_insert(new_formula,
                                        insert['raw'],
                                        insert['reference'])

            # Otherwise, raise an error
            else:
                raise ValueError('Formula missing necessary keys')

        return None