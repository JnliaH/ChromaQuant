#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

COPYRIGHT STATEMENT:

ChromaQuant – A quantification software for complex gas chromatographic data

Copyright (c) 2025, by Julia Hancock
              Affiliation: Dr. Julie Elaine Rorrer
	      URL: https://www.rorrerlab.com/

License: BSD 3-Clause License

---

SUBPACKAGE FOR EXPORTING QUANTIFICATION RESULTS

Julia Hancock
Started 10-27-25

"""
import os
import pandas as pd
from datetime import datetime
import openpyxl as xl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment

""" FUNCTIONS """

def exportToExcel(sname,sinfo,output_directory,analysis_config,liqDF=[],gasFIDDF=[],gasTCDDF=[]):

    def getExcelPath(sname,path):

        #Function for checking if file exists and adding number if so, used for new breakdowns
        def checkExcel(path):
            
            #Inspired by https://stackoverflow.com/questions/13852700/create-file-but-if-name-exists-add-number
            
            filename, extension = os.path.splitext(path)
            i = 1
            
            while os.path.exists(path):
                path = filename + " ("+str(i)+")" + extension
                i += 1
            
            return path

        #Get current datetime string
        nows = datetime.now().strftime('%Y%m%d')

        #Define breakdown file name
        bfn = sname+"_Breakdown_"+nows+".xlsx"

        output_path = checkExcel(os.path.join(path,bfn))

        return output_path

    def exportSampleInfo(eWriter,sinfo,append_dict,sheet_name):

        # Get dataframe for sample info
        sinfo_DF = pd.DataFrame(sinfo,index=[0])

        # For every entry in the append dictionary..
        for key in append_dict:

            # Add the key-value pair to sinfo_DF
            sinfo_DF[key] = append_dict[key]
        
        # Export sample info to specified sheet
        sinfo_DF.to_excel(eWriter, sheet_name=sheet_name,startcol=1, startrow=1, index=False)

        return None

    def exportRawData(eWriter,BreakdownDF,sheet_name):

        #Export raw data dataframe to Excel
        BreakdownDF.to_excel(eWriter, sheet_name=sheet_name,startcol=1, startrow=4, index=False)

        return None

    def getSinfoCell(sh,search_term):

        # Define the sample info header row
        sinfo_row_header = 2

        # Define the sample info value row
        sinfo_row_value = 3

        # Predefine default sinfo mass cell
        cell_location = '$A$1'

        # For every cell in the row
        for cell in sh[sinfo_row_header]:
            
            # If the cell header is the internal standard name value..
            if cell.value == search_term:

                # Get the cell column
                sinfo_column = cell.column_letter

                # Construct the IS_cell_mass cell
                cell_location = f'${sinfo_column}${sinfo_row_value}'

                return cell_location
            
            # Otherwise, pass
            else:

                pass

        return cell_location
    
    # Function to get the cell where the internal standard area is
    def getISAreaCell(sh,IS_name):

        # Define the compound name column string
        column_name = 'E'

        # Define the FID area column string
        column_area = 'C'

        # Get all cells in the compound name column
        column_cells = sh[column_name]

        # Predefine the cell row variable with a default
        cell_row = 1

        # Define the cell coordinate for the internal standard area as a default
        cell_coordinate = f'${column_area}${cell_row}'

        # For every cell in the column...
        for cell in column_cells:

            # If the cell's value is equal to the internal standard name, return the row number
            if cell.value == IS_name:

                # Extract the cell's row
                cell_row = cell.row

                # Redefine the cell coordinate
                cell_coordinate = f'${column_area}${cell_row}'

                return [cell_row,cell_coordinate]

            else:
                pass

        if cell_row == 0:
            print('[outputQuant] ERROR - could not find internal standard name in exported Excel')
        
        else:
            pass
        
        return [cell_row,cell_coordinate]

    # Function to export the formula to calculate final gas volume
    def exportGasVolumeFormula(sh):
        
        # Get the cell locations for each relevant variable and the output
        initial_pressure = getSinfoCell(sh,'Quench Pressure (psi)')
        final_pressure = getSinfoCell(sh,'Gas Bag Pressure (psia)')
        initial_volume = getSinfoCell(sh,'Reactor Volume (mL)')
        final_volume = getSinfoCell(sh,'Gas Bag Volume (m^3)')
        
        # Set the output formula
        sh[final_volume] = f'={initial_pressure}*{initial_volume}/{final_pressure}*1e-6'
    
        return final_volume
    
    # Function to export formulas 
    def exportFormulas(sh,DF_length,mass_label_list,formula_dict,header_dict):

        # Unpack the mass and label column names from mass_label_list
        mass_column = mass_label_list[0]
        label_column = mass_label_list[1]
        
        def exportHeaders(sh,header_dict,row):

            # For each entry in header_dict..
            for column in header_dict:

                # Export the header
                sh[f'{column}{row}'] = header_dict[column]

            return None
    
        # Function to simulate dragging down in Excel
        def dragDown(sh,formula_string,out_column,start_row,end_row,skip_row):

            # Loop through every row in the range provided
            for row in range(start_row, end_row):

                if row != skip_row:
                    # Define the output formula by formatting the provided string
                    formula = formula_string.format(row=row,column=out_column)

                    # Enter the formula as formatted to the current cell
                    # NOTE: This assumes the row of the cell being used is the same
                    #       as the row of the exported formula
                    sh[f'{out_column}{row}'] = formula
                
                else:
                    pass
                
            return None
        
        # Export parameters
        start_row = 6
        end_row = start_row + DF_length

        # Export headers
        exportHeaders(sh,header_dict,start_row-1)

        # Loop through all formulas
        for column in formula_dict:

            # Export each formula along the specified column across the specified row range
            dragDown(sh,formula_dict[column],column,start_row,end_row,IS_cell_area_row)

        # Define the formula_details variable to contain row and column information relevant for the distribution matrix
        formula_details = [start_row,end_row,mass_column,label_column]

        return formula_details

    # Function to export a distribution matrix
    def exportMatrix(sh,liqDF,analysis_config,cell_start,formula_details):

        # Define some colors for later use
        color0 = 'ff92cddc'
        color1 = 'ffb7dee8'
        color2 = 'ffdaeef3'

        # Define pattern fill objects using these colors
        pattern0 = PatternFill(start_color=color0, end_color=color0, fill_type='solid')
        pattern1 = PatternFill(start_color=color1, end_color=color1, fill_type='solid')
        pattern2 = PatternFill(start_color=color2, end_color=color2, fill_type='solid')

        # Define the centered alignment object
        center_alignment = Alignment(horizontal='center',vertical='center')

        # Function to export the matrix headers
        def matrixHeaders(sh,row_start,column_start,CN_max,CT_length,CT_list,CT_keylist,format_list):

            # Extract the list of patterns
            pattern_list = format_list[0]

            # Extract the center alignment object
            center_alignment = format_list[1]

            # Get the index to use to iterate through the list of compound types as the loop progresses
            i = 0

            # Set the value of the starting cell to 'Carbon Number'
            sh[cell_start] = 'Carbon Number'

            # Set the background color
            sh[cell_start].fill = pattern_list[0]

            # Align the cell
            sh[cell_start].alignment = center_alignment
            
            # Define the range over which to merge cells for carbon number label
            header_merge_range = f'{cell_start}:{sh.cell(row=row_start+1,column=column_start).coordinate}'
            sh.merge_cells(header_merge_range)

            # Export the top headers starting at one column after cell_start
            for column in range(column_start+1,column_start+CT_length+1,1):

                # Export value to first header
                sh.cell(row=row_start,column=column).value = CT_list[i]

                # Export key to subheader
                sh.cell(row=row_start+1,column=column).value = CT_keylist[i]

                # Set the background colors
                sh.cell(row=row_start,column=column).fill = pattern_list[0]
                sh.cell(row=row_start+1,column=column).fill = pattern_list[1]

                # Align the cell
                sh.cell(row=row_start,column=column).alignment = center_alignment
                sh.cell(row=row_start+1,column=column).alignment = center_alignment

                # Increase CT_list index by one
                i += 1

            # Get the starting carbon number for exporting
            CN = 1

            # Export the side headers for the carbon number up to the maximum
            for row in range(row_start+2,row_start+CN_max+2,1):

                # Export value
                sh.cell(row=row,column=column_start).value = CN

                # Set the background color
                sh.cell(row=row,column=column_start).fill = pattern_list[1]

                # Align the cell
                sh.cell(row=row,column=column_start).alignment = center_alignment

                # Increase the carbon number by one
                CN += 1
            
            return None
        
        def acquireRanges(formula_details):

            # Unpack the formula_details object
            start_row = formula_details[0]
            end_row = formula_details[1]
            mass_column = formula_details[2]
            label_column = formula_details[3]

            # Define the mass and label ranges
            mass_range = f'${mass_column}${start_row}:${mass_column}${end_row}'
            label_range = f'${label_column}${start_row}:${label_column}${end_row}'

            return [mass_range,label_range]
        
        def matrixFormulas(sh,row_start,column_start,CN_max,CT_length,range_list,format_list):

            # Extract the list of patterns
            pattern_list = format_list[0]

            # Extract the center alignment object
            center_alignment = format_list[1]

            # Extract the list of ranges
            mass_range = range_list[0]
            label_range = range_list[1]

            # Get the row and column for the relevant headers to be concatenated
            row_CT_header = row_start + 1
            column_CN_header = column_start

            # Get the row and column for the first cell in the formula block
            row_start += 2
            column_start += 1

            # For every cell in the formula block, set the formula to sum every mass with a label string that matches the current cell
            for column in range(column_start,column_start+CT_length,1):

                for row in range(row_start,row_start+CN_max,1):

                    sh.cell(row=row,column=column).value = f'=SUMIF({label_range},CONCATENATE({get_column_letter(column)}${row_CT_header},${get_column_letter(column_CN_header)}{row}),{mass_range})'


            return None
        
        # Get maximum carbon number in the DataFrame
        CN_max = int(liqDF['Carbon Number'].max())

        # Get the dictionary of compound types from analysis_config
        CT_dict = analysis_config['CT_Dict']

        # Get lists of keys and values from the dictionary, flattening them into a list
        CT_keylist = list(CT_dict.keys())
        CT_list = list(CT_dict.values())

        # Get the total number of unique compound types from CT_list
        CT_length = len(CT_list)

        # Get the starting cell
        cell = sh[cell_start]

        # Get the column and row from the starting cell
        row_start = cell.row
        column_start = cell.column
        
        # Run the matrix headers function
        matrixHeaders(sh,row_start,column_start,CN_max,CT_length,CT_list,CT_keylist,[[pattern0,pattern1,pattern2],center_alignment])

        # Run the acquire ranges function
        range_list = acquireRanges(formula_details)

        # Run the matrix formulas function
        matrixFormulas(sh,row_start,column_start,CN_max,CT_length,range_list,[[pattern0,pattern1,pattern2],center_alignment])

        return None

    
    # Get the output Excel path
    output_path = getExcelPath(sname,output_directory)

    # Define dictionaries to append to existing sample info
    liq_append_dict = {}

    gas_append_dict = {'Gas Bag Temperature (K)':analysis_config['sample-injection-conditions']['gas-bag-temp-C']+273.15,\
                          'Gas Bag Pressure (psia)':analysis_config['sample-injection-conditions']['gas-bag-pressure-psia'],\
                          'Ideal Gas Constant (m^3*Pa/K*mol)':8.314,\
                          'Gas Bag Volume (m^3)':''}

    # Open the Pandas Excel writer object
    with pd.ExcelWriter(output_path,engine='xlsxwriter') as eWriter:

        #If liquids data are passed...
        if isinstance(liqDF,pd.DataFrame):

            # Export sample info to liquids sheet
            exportSampleInfo(eWriter,sinfo,liq_append_dict,'Liquid FID')

            # Export liquid data and formulas
            exportRawData(eWriter,liqDF,'Liquid FID')

        #If gas data are passed...
        if isinstance(gasFIDDF, pd.DataFrame) and isinstance(gasTCDDF, pd.DataFrame):
            
            # Export sample info to each sheet
            exportSampleInfo(eWriter,sinfo,gas_append_dict,'Gas FID')
            exportSampleInfo(eWriter,sinfo,gas_append_dict,'Gas TCD')

            # Export raw data to each sheet
            exportRawData(eWriter,gasFIDDF,'Gas FID')
            exportRawData(eWriter,gasTCDDF,'Gas TCD')

    # Reopen the Excel sheet with a non-pandas Excel writer object
    try:

        # Open the workbook
        workbook = xl.load_workbook(filename=output_path)
        print("[outputQuant] Reopened workbook with openpyxl successfully")

        # If liquids data are passed...
        if isinstance(liqDF,pd.DataFrame):
            
            # Open the liquid FID spreadsheet
            sheet = workbook['Liquid FID']

            # Get the name of internal standard
            IS_name = sinfo['Internal Standard Name']

            # Get the internal standard mass location 
            IS_cell_mass = getSinfoCell(sheet,'Internal Standard Mass (mg)')
        
            # Get the internal standard area location
            IS_cell_area_row, IS_cell_area = getISAreaCell(sheet,IS_name)

            # Dictionary for liquid headers to be exported
            liq_header_dict = {'N':'Area Ratio','O':'Mass (mg)','P':'Label String'}

            # Dictionary for all liquid formulas to be exported
            liq_formula_dict = {'N':'=C{row}/'+f'{IS_cell_area}','O':f'={IS_cell_mass}'+'*N{row}/L{row}','P':'=CONCATENATE(I{row},K{row})'}

            # Export the formulas for liquids
            formula_details = exportFormulas(sheet,len(liqDF),['O','P'],liq_formula_dict,liq_header_dict)

            # Define the starting cell for the distribution matrix
            cell_start = 'R5'

            # Export the distribution matrix
            exportMatrix(sheet,liqDF,analysis_config,cell_start,formula_details)
            

        #If gas data are passed...
        if isinstance(gasFIDDF, pd.DataFrame) and isinstance(gasTCDDF, pd.DataFrame):

            ## GAS FID
            # Open the gas FID spreadsheet
            sheet = workbook['Gas FID']

            # Get the name of the internal standard
            IS_name = 'CO2'

            # Get the environmental conditions locations
            gas_bag_temp = getSinfoCell(sheet,'Gas Bag Temperature (K)')
            gas_bag_pressure = getSinfoCell(sheet,'Gas Bag Pressure (psia)')
            gas_constant = getSinfoCell(sheet,'Ideal Gas Constant (m^3*Pa/K*mol)')

            # Export the gas volume formula and get its location
            gas_volume = exportGasVolumeFormula(sheet)

            # Dictionary for gas FID headers to be exported
            gasFID_header_dict = {'J':'Carbon Number','K':'RF (Area/vol.%)','L':'RF Source',\
                                  'M':'MW (g/mol)','N':'Vol.%','O':'Moles (mol)','P':'Mass (mg)','Q':'Label String'}

            # Dictionary for all gas FID formulas to be exported
            gasFID_formula_dict = {'N':'=$C{row}/$K{row}',\
                                   'O':f'={gas_bag_pressure}*6894.76'+'*$N{row}/100'+f'*{gas_volume}/({gas_constant}*{gas_bag_temp})',\
                                   'P':'=$O{row}*$M{row}*1000',\
                                   'Q':'=CONCATENATE(I{row},J{row})'}

            # Export the formulas for gas FID
            formula_details = exportFormulas(sheet,len(gasFIDDF),['P','Q'],gasFID_formula_dict,gasFID_header_dict)

            # Define the starting cell for the distribution matrix
            cell_start = 'S5'

            # Export the distribution matrix
            exportMatrix(sheet,gasFIDDF,analysis_config,cell_start,formula_details)

            ## GAS TCD
            # Open the gas TCD spreadsheet
            sheet = workbook['Gas TCD']

            # Get the internal standard volume location
            IS_cell_volume = getSinfoCell(sheet,'Injected CO2 (mL)')

            # Get the name of the internal standard
            IS_name = 'CO2'

            # Get the environmental conditions locations
            gas_bag_temp = getSinfoCell(sheet,'Gas Bag Temperature (K)')
            gas_bag_pressure = getSinfoCell(sheet,'Gas Bag Pressure (psia)')
            gas_constant = getSinfoCell(sheet,'Ideal Gas Constant (m^3*Pa/K*mol)')

            # Export the gas volume formula and get its location
            gas_volume = exportGasVolumeFormula(sheet)

            # Dictionary for gas TCD headers to be exported
            gasTCD_header_dict = {}


        #Save the workbook
        workbook.save(filename=output_path)

    # Raise exceptions if necessary
    except FileNotFoundError:
        print("[outputQuant] Could not reopen workbook for analysis, only exported raw data")

    except Exception as e:
        print(f"An error occurred: {e}")

    return None