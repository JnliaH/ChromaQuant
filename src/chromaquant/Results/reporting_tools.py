#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

This submodule contains functions used in Excel reporting, particularly
by the Results class (see Results).

"""

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from pandas import ExcelWriter
from ..data import Breakdown, Table, Value
from ..data.dataset import DataSet

""" FUNCTIONS """


# Function to write a Breakdown to Excel
def report_breakdown(breakdown: Breakdown,
                     writer: ExcelWriter):
    """
    Writes a Pandas Breakdown to Excel using passed writer.

    Parameters
    ----------
    breakdown : Breakdown
        Pandas Breakdown to export.
    writer : ExcelWriter
        Pandas ExcelWriter used in exporting.

    Returns
    ----------
    None

    """
    # Get the start row based on whether there is a header or not
    start_row = \
        breakdown.start_row if breakdown.header == '' \
        else breakdown.start_row + 1

    # Write the passed DataFrame to passed path
    breakdown.data.to_excel(writer,
                            sheet_name=breakdown.sheet,
                            startcol=breakdown.start_column,
                            startrow=start_row,
                            index=False)


# Function to write a header to Excel
def report_header(dataset: DataSet,
                  workbook: Workbook):
    """
    Writes a DataSet's header to Excel.

    Parameters
    ----------
    dataset : DataSet
        Dataset to export the header of.
    workbook : Workbook
        Active openpyxl workbook.

    Returns
    -------
    None

    """

    # If the dataset has a blank header...
    if dataset.header == '':
        # Don't report header
        return None

    # If dataset's sheet does not exist in workbook...
    if dataset.sheet not in workbook.sheetnames:
        # Create it
        workbook.create_sheet(dataset.sheet)

    # Open the dataset's sheet
    sheet = workbook[dataset.sheet]

    # Write the header to the start cell (adjusted from absolute)
    cell = sheet.cell(dataset.start_row + 1,
                      dataset.start_column + 1,
                      value=dataset.header)

    # Center the header cell
    cell.alignment = Alignment(horizontal='center')

    # Try to get the number of columns in the DataSet's data,
    # only working for Table, Breakdown
    try:
        num_cols = dataset.data.shape[1]

        # If this works, merge the header cell with adjacent cells
        # to center the header across the DataFrame
        sheet.merge_cells(start_row=dataset.start_row + 1,
                          start_column=dataset.start_column + 1,
                          end_row=dataset.start_row + 1,
                          end_column=dataset.start_column + num_cols)

    # If this does not work, pass
    except AttributeError:
        pass

    return None


# Function to write a Table to Excel
def report_table(table: Table,
                 writer: ExcelWriter):
    """
    Writes a Table to Excel.

    Parameters
    ----------
    table: Table
        Table to export.
    writer : ExcelWriter
        Pandas ExcelWriter used in exporting.

    Returns
    -------
    None

    """

    # Get the start row based on whether there is a header or not
    start_row = \
        table.start_row if table.header == '' \
        else table.start_row + 1

    # Write the passed DataFrame to passed path
    table.data.to_excel(writer,
                        sheet_name=table.sheet,
                        startcol=table.start_column,
                        startrow=start_row,
                        index=False)

    return None


# Function to write a Value to Excel
def report_value(value: Value,
                 workbook: Workbook):
    """
    Writes a Value to Excel.

    Parameters
    ----------
    value: Value
        Value to export.
    workbook : Workbook
        Active openpyxl workbook.

    Returns
    -------
    None

    """

    # If Value's sheet does not exist in workbook...
    if value.sheet not in workbook.sheetnames:
        # Create it
        workbook.create_sheet(value.sheet)

    # Open the Value's sheet
    sheet = workbook[value.sheet]

    # Get the start row based on whether there is a header or not
    start_row = \
        value.start_row + 1 if value.header == '' \
        else value.start_row + 2

    # Write the Value's data to the cell below, adjusting from absolute
    sheet.cell(start_row,
               value.start_column + 1,
               value=value.data)

    return None


# Function to remove a Table or Breakdown's column header borders
def remove_borders(table_or_breakdown: Table | Breakdown,
                   workbook: Workbook):
    """
    Removes borders around a Table or Breakdown's column headers.

    Parameters
    ----------
    table_or_breakdown : Table | Breakdown
        Table or Breakdown instance.
    workbook : Workbook
        Active openpyxl workbook.

    Returns
    -------
    None

    """

    # Open the dataset's sheet
    sheet = workbook[table_or_breakdown.sheet]

    # If the dataset has a header...
    if table_or_breakdown.header != '':

        # Set the starting row to the start_row plus 2
        start_row = table_or_breakdown.start_row + 2

    # Otherwise, set the starting row to the start_row plus 1
    else:
        start_row = table_or_breakdown.start_row + 1

    # For every column in the index...
    for column_index in range(
       table_or_breakdown.start_column,
       table_or_breakdown.start_column +
       len(table_or_breakdown.data.columns) + 1):

        # Get the current cell
        cell = sheet.cell(start_row,
                          column_index + 1)

        # Set the cell's border to None
        cell.border = None

    return None


# Function to set all cells in a workbook to a default format
def set_default_format(workbook: Workbook):
    """
    Sets all cells in an active openpyxl workbook to a default format.

    Parameters
    ----------
    workbook : Workbook
        Active openpyxl workbook.

    Returns
    -------
    None

    """

    # For every worksheet in the workbook...
    for sheet in workbook.worksheets:
        # For every column in the sheet...
        for column in sheet.iter_cols():
            # Get the maximum length of all values in the column
            column_width = max(len(str(cell.value)) for cell in column)
            # If the column_width is above 25, set to 25
            column_width = 25 if column_width > 25 else column_width
            # If the column_width is less than 8, set to 8
            column_width = 8 if column_width < 8 else column_width
            # Get the column's index
            col_index = column[0].column
            # Set the column's width to the minimum required to fit all values
            sheet.column_dimensions[get_column_letter(col_index)].width = \
                column_width

    return None
