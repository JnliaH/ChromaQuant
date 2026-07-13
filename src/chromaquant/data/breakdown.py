#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

The Breakdown class is used to conditionally aggregate data from
a Table instance. It enables users to construct 1D and 2D DataFrames
that can then be used in reporting or as desired.

"""

from __future__ import annotations

import logging
from openpyxl.utils import get_column_letter
from pandas import DataFrame
import pandas as pd
from typing import Any, TYPE_CHECKING
if TYPE_CHECKING:
    from ..results import Results
from .dataset import DataSet
from ._column_id import _ColumnID
from .table import Table
from ..logging_and_handling import setup_logger, setup_error_logging

""" LOGGING AND HANDLING """

# Create a logger
logger = logging.getLogger(__name__)

# Format the logger
logger = setup_logger(logger)

# Get an error logging decorator
error_logging = setup_error_logging(logger)

""" CLASS"""


# Class definition for Breakdown
class Breakdown(DataSet):
    """
    Class used to conditionally aggregate data from a Table

    Parameters
    ----------
    start_cell : str, optional
        Reference to cell in Excel where data will be reported,
        referring to the top-left of report range. Must be a
        valid Excel cell (e.g., 'A1', '$B$2').

    sheet : str, optional
        Name of Excel worksheet (sheet within workbook) where data will
        be reported.

    conditional_aggregate: str, optional
        String representation of the desired conditional aggregate formula.
        Options area 'SUMIFS', 'COUNTIFS', 'AVERAGEIFS', 'MINIFS', and
        'MAXIFS'. Default is 'SUMIFS'.

    header: str, optional
        Header to add above a dataset, equivalent to a title.

    results: Results, optional
        Results object that mediates this DataSet.

    Raises
    ------
    ValueError
        If a conditional_aggregate is passed that is not a valid option.
    ValueError
        If sheet is set to a blank string.
    ValueError
        If start_cell is set to an invalid Excel cell.
    ValueError
        If no summarize_column is provided for conditional aggregates
        besides COUNTIFS when creating a conditional aggregate formula.
    ValueError
        If more than two key-value pairs are in groups_to_summarize or
        a key does not match a passed group by column when creating a
        2D breakdown.

    """

    # Initialize
    def __init__(self,
                 start_cell: str = '',
                 sheet: str = '',
                 conditional_aggregate: str = 'SUMIFS',
                 header: str = '',
                 results: Results = None):

        # Run DataSet initialization
        super().__init__(data=DataFrame(),
                         start_cell=start_cell,
                         sheet=sheet,
                         type='Breakdown',
                         header=header,
                         results=results)

        # Create attributes
        self._data = DataFrame()
        self.allowed_conditional_aggregates = ['SUMIFS',
                                               'COUNTIFS',
                                               'AVERAGEIFS',
                                               'MINIFS',
                                               'MAXIFS']
        self._breakdown_cache = {}
        self.length = len(self._data)
        self.columns: list[str] = []

        # If the conditional aggregate is not equal to an expected value...
        if conditional_aggregate not in self.allowed_conditional_aggregates:
            # Raise an error
            raise ValueError(('Conditional aggregate not recognized: '
                              f'{conditional_aggregate}'))
        # Otherwise, assign to an attribute
        else:
            self.conditional_aggregate = conditional_aggregate

        # Update the Breakdown
        self._update_breakdown()

    """ PROPERTIES """
    # Data properties - GETTER ONLY
    # Getter
    @property
    def data(self) -> Any:
        """
        Get, set, or delete data stored in the DataSet. Common types include
        str, bool, int, float, list, dict, or pandas DataFrame.
        """
        return self._data

    # Setter
    @data.setter
    def data(self, value: DataFrame | None):
        self._data = value

    # Deleter
    @data.deleter
    def data(self):
        del self._data
        self._data = DataFrame()

    # Footprint
    # Only getter
    @property
    def footprint(self):

        # Create a default footprint
        self._footprint = {'header': '',
                           'subheader': '',
                           'body': {}}

        # Update the breakdown
        self._update_breakdown()

        # For every column in columns...
        for column in self.columns:

            # Get the start cell's column letter, adjusting from absolute
            col_letter = \
                get_column_letter(self.start_column + 1 +
                                  self._data.columns.get_loc(column))

            # If there is a header...
            if self.header != '':

                # Get a start row, adjusting from absolute
                start_row = self.start_row + 3

                # Get an end row, adjusting from absolute
                end_row = self.start_row + 2 + self.length

                # Get a range reference, adjusting from absolute
                column_range = (f"${col_letter}${start_row}:"
                                f"${col_letter}${end_row}")

            # If there isn't a header...
            else:
                # Get a start row, adjusting from absolute
                start_row = self.start_row + 2

                # Get an end row, adjusting from absolute
                end_row = self.start_row + 1 + self.length

                # Get a range reference, adjusting from absolute
                column_range = (f"${col_letter}${start_row}:"
                                f"${col_letter}${end_row}")

            # Add the column range to the body dictionary in footprint
            self._footprint['body'][column] = column_range

        # Get the starting row and column indices
        start_column_index, start_row = self.get_cell_indices(self.start_cell)

        # Get the starting column's letter
        start_column = get_column_letter(start_column_index + 1)

        # Find the final column index by adding the length of the columns
        # attribute
        end_column_index = start_column_index + len(self.columns)

        # Get the final column letter
        end_column = get_column_letter(end_column_index)

        # If there is a header...
        if self.header:

            # Add the header range
            self._footprint['header'] = \
                f'${start_column}${start_row+1}:${end_column}${start_row+1}'

            # Add the subheader range
            self._footprint['subheader'] = \
                f'${start_column}${start_row+2}:${end_column}${start_row+2}'

        # Otherwise...
        else:

            # Add the subheader range
            self._footprint['subheader'] = \
                f'${start_column}${start_row+1}:${end_column}${start_row+1}'

        return self._footprint

    # Sheet properties
    # Getter
    @property
    def sheet(self) -> str:
        """
        Get, set, or delete the name of the Excel worksheet to report to.
        """
        return self._sheet

    # Setter
    @sheet.setter
    def sheet(self, value: str):
        if value == '':
            raise ValueError('Breakdown sheet cannot be an empty string.')
        self._sheet = value
        self._update_breakdown()
        if self._mediator is not None:
            self._mediator.update_datasets()

    # Deleter
    @sheet.deleter
    def sheet(self):
        self._sheet = 'Sheet1'
        self._update_breakdown()

    # Start cell properties
    # Getter
    @property
    def start_cell(self) -> str:
        """
        Get or set the Excel reference where data will be reported.
        """
        return self._start_cell

    # Setter
    @start_cell.setter
    def start_cell(self, value: str):
        # Try...
        try:
            # Get the cell's absolute indices
            self.start_column, self.start_row = self.get_cell_indices(value)
            # Set the starting cell
            self._start_cell = value
        # If an exception occurs...
        except Exception as e:
            raise ValueError(f'Passed start cell is not valid: {e}')
        self._update_breakdown()
        if self._mediator is not None:
            self._mediator.update_datasets()

    # Deleter
    @start_cell.deleter
    def start_cell(self):
        # Reset the starting cell
        self._start_cell = '$A$1'
        # Get the cell's absolute indices
        self.start_column, self.start_row = self.get_cell_indices('$A$1')
        self._update_breakdown()

    """ METHODS """
    # Method to get a given column's id
    @error_logging
    def column_id(self, column_name: str):
        """
        Method that returns a column's ID dictionary.

        Parameters
        ----------
        column_name : str
            Name of a column of interest.

        Returns
        -------
        new_column_id : dict[str, str]
            ColumnID of interest.
        """

        # Create a ColumnID
        new_column_id = _ColumnID(self, column_name)

        return new_column_id

    # Method to create a conditional aggregate formula based on criteria
    def _create_conditional_aggregate_formula(self,
                                              table: Table,
                                              criteria: dict[str, str],
                                              summarize_column: str = ''
                                              ) -> str:
        """
        Method that creates a conditional aggregate formula by referencing
        passed Table and criteria.

        Parameters
        ----------
        table : Table
            An instance of Table.

        criteria : dict[str, str]
            A dictionary containing one or two key-value pairs where the keys
            are ranges in the DataFrame to aggregate and the values are cell
            references used in the conditional statement.

        summarize_column : str, optional
            Name of the column to summarize. Optional for COUNTIFS but required
            for other conditional aggregate formulas, by default ''.

        Returns
        -------
        formula: str
            The resulting conditional aggregate formula.

        Raises
        ------
        ValueError
            If no summarize_column is provided for conditional aggregates
            besides COUNTIFS.

        """

        # Create a blank formula template
        formula_template = ''

        # If the conditional aggregate is COUNTIFS...
        if self.conditional_aggregate == 'COUNTIFS':

            # Create a COUNTIFS-specific inner formula template
            formula_template = ''

        # Otherwise, if summarize column is not empty...
        elif summarize_column != '':
            # Create an inner formula template
            formula_template = f"{table.reference[summarize_column]['range']}"

        # Otherwise, raise an error
        else:
            raise ValueError(('Argument summarize_column cannot be blank for'
                              'conditional aggregates besides COUNTIFS'))

        # For every key-value pair in criteria...
        for cell_or_range in criteria:

            # Add criteria range to formula_template
            formula_template += f', {cell_or_range}'
            # Add criteria to formula_template
            formula_template += f', {criteria[cell_or_range]}'

        # Get the final formula by wrapping in the passed conditional aggregate
        formula = f'={self.conditional_aggregate}({formula_template})'

        return formula

    # Method to create a one-dimensional breakdown
    def create_1D(self,
                  table: Table,
                  group_by_column: str,
                  summarize_column: str,
                  groups_to_summarize: list[str] = None):
        """
        Method to create a one-dimensional breakdown.

        Parameters
        ----------
        table : Table
            An instance of Table.

        group_by_column : str
            The name of the column by which to aggregate results.

        summarize_column : str
            The name of the column to summarize.

        groups_to_summarize : list[str], optional
            An optional list of groups to include in the breakdown,
            overwriting the default list created that includes all
            groups present in the current Table, by default None.

        Returns
        -------
        None

        """

        # If groups_to_summarize is not none...
        if groups_to_summarize is not None:
            # Set unique_groups to a shallow copy of groups_to_summarize
            unique_groups = groups_to_summarize[:]

        # Otherwise...
        else:
            # Get a list of unique values in Table in the group by column
            # NOTE: IGNORES NONE
            unique_groups = [group for group in
                             table.data[group_by_column].unique()
                             if group is not None]

            # Try to sort the list
            try:
                unique_groups.sort()

            # If a type error occurs, pass
            except TypeError:
                pass

        # Create an empty 1D DataFrame using the unique groups
        self._data = DataFrame({group: [None] for group in unique_groups})

        # Create an index for the current header cell
        header_cell_index = 0

        # Get the start cell's absolute indices
        absolute_start_col, absolute_start_row = \
            self.get_cell_indices(self.start_cell)

        # For every group...
        for group in unique_groups:

            # Get the current header's column and row
            start_col = \
                get_column_letter(absolute_start_col + 1 + header_cell_index)
            start_row = \
                absolute_start_row + 1

            # Get the current header cell
            header_cell = f'{start_col}${start_row}'

            # Define the criteria dictionary
            criteria = {table.reference[group_by_column]['range']: header_cell}

            # Get a formula string
            formula_string = \
                self._create_conditional_aggregate_formula(table,
                                                           criteria,
                                                           summarize_column
                                                           )

            # Add the formula string to the current group entry
            self._data[group] = formula_string

            # Iterate the header cell index
            header_cell_index += 1

        # Set the data length
        self.length = len(self._data)

        # Set the breakdown cache
        self._breakdown_cache = {'function': self.create_1D,
                                 'arguments': {'table':
                                               table,
                                               'group_by_column':
                                               group_by_column,
                                               'summarize_column':
                                               summarize_column,
                                               'groups_to_summarize':
                                               groups_to_summarize}}

        return None

    # Method to create a two-dimensional breakdown
    def create_2D(self,
                  table: Table,
                  group_by_col_1: str,
                  group_by_col_2: str,
                  summarize_column: str,
                  groups_to_summarize: dict[str, list[str]] = None):
        """
        Method to create a two-dimensional breakdown.

        Parameters
        ----------
        table : Table
            An instance of Table.

        group_by_col_1 : str
            The name of the first column by which to aggregate results.

        group_by_col_2 : str
            The name of the second column by which to aggregate results.

        summarize_column : str
            The name of the column to summarize.

        groups_to_summarize : dict[str, list[str]], optional
            An optional dictionary of groups to include in the breakdown.
            The key in each pair is the name of a passed group-by column
            and the value is a list of groups to include in the breakdown.
            This overwrites the default list created that includes all
            groups present in the current Table under the passed group-by
            columns, by default None.

        Returns
        -------
        ValueError
            If more than two key-value pairs are in groups_to_summarize or
            a key does not match a passed group by column when creating a
            2D breakdown.

        """

        # Create an empty dictionary to contain lists of unique groups
        unique_groups = {group_by_col_1: [], group_by_col_2: []}

        # If groups_to_summarize is not none...
        if groups_to_summarize is not None:

            # Get a list of booleans based on whether each key
            # in groups_to_summarize matches a group_by_col
            check_groups_list = [True if group_col in
                                 [group_by_col_1, group_by_col_2]
                                 else False for group_col in
                                 groups_to_summarize]

            # If groups to summarize is longer than two or contains a string
            # that is not equal to either column string...
            if len(groups_to_summarize) > 2 or not all(check_groups_list):

                # Raise an error
                raise ValueError('Unexpected keys in groups_to_summarize.')

            # Otherwise, pass
            else:
                pass

        # Otherwise, create an empty dictionary
        else:
            groups_to_summarize = {}

        # For every entry in unique_groups...
        for group_col in unique_groups:

            # If the group column is present in groups_to_summarize...
            if group_col in groups_to_summarize:

                # Set unique_groups[group_col] to a shallow
                # copy of groups_to_summarize entry
                unique_groups[group_col] = groups_to_summarize[group_col][:]

            # Otherwise...
            else:
                # Set unique_groups[group_col] to a list of unique values
                # in the group by column
                unique_groups[group_col] = [group for group in
                                            table.data[group_col].unique()
                                            if group is not None]

                # Try to sort the list
                try:
                    unique_groups[group_col].sort()

                # If a type error occurs, pass
                except TypeError:
                    pass

        # Create a new column for group_2 (row headers)
        self._data[group_by_col_2] = \
            [group for group in unique_groups[group_by_col_2]]

        # Get the start cell's absolute indices
        absolute_start_col, absolute_start_row = \
            self.get_cell_indices(self.start_cell)

        # Create column index to track current column
        column_index = 0

        # For every group in unique_groups group_by_1 (treat as columns)...
        for group_1 in unique_groups[group_by_col_1]:

            # Get the current column header's column and row
            column_start_col = \
                get_column_letter(absolute_start_col + 2 +
                                  column_index)
            column_start_row = \
                absolute_start_row + 2

            # Get the current column header cell
            column_header_cell = f'{column_start_col}${column_start_row}'

            # Create row index to track current row
            row_index = 0

            # For every group in unique_groups group_by_2 (treat as rows)...
            for group_2 in unique_groups[group_by_col_2]:

                # Get the current row header's column and row
                row_start_col = \
                    get_column_letter(absolute_start_col + 1)
                row_start_row = \
                    absolute_start_row + 3 + row_index

                # Get the current row header cell
                row_header_cell = f'${row_start_col}{row_start_row}'

                # Define the criteria dictionary
                criteria = {table.reference[group_by_col_1]['range']:
                            column_header_cell,
                            table.reference[group_by_col_2]['range']:
                            row_header_cell}

                # Get a formula string
                formula_string = \
                    self._create_conditional_aggregate_formula(table,
                                                               criteria,
                                                               summarize_column
                                                               )

                # Add the formula string to the current group entry
                self._data.at[row_index, group_1] = formula_string

                # Iterate the row cell index
                row_index += 1

            # Iterate the header cell index
            column_index += 1

        # Set the data length
        self.length = len(self._data)

        # Set the breakdown cache
        self._breakdown_cache = {'function': self.create_2D,
                                 'arguments': {'table':
                                               table,
                                               'group_by_col_1':
                                               group_by_col_1,
                                               'group_by_col_2':
                                               group_by_col_2,
                                               'summarize_column':
                                               summarize_column,
                                               'groups_to_summarize':
                                               groups_to_summarize}}

        return None

    # Method to create a Breakdown by merging passed Breakdowns
    @error_logging
    def merge_breakdowns(self,
                         breakdown_list: list[DataSet]):

        # Get a set of all DataSet shapes
        shapes = {dataset.data.shape for dataset in breakdown_list}

        # If all DataSets are of the same dimension...
        if len(shapes) == 1:

            # Create an empty DataFrame using the shape of the first Breakdown
            self._data = \
                pd.DataFrame().reindex_like(
                    breakdown_list[0]._data
                    ).replace({float('nan'): None})

            # Set the first column values to the first column of the first
            # passed DataFrame
            self._data[breakdown_list[0]._data.columns[0]] = \
                breakdown_list[0]._data[breakdown_list[0]._data.columns[0]]

            # Get an initial row index
            row_index = 2 if not self._header else 3

            # For every cell in the DataFrame...
            for index, row in self._data.iterrows():
                # Get an initial column index
                column_index = 2
                for column in self._data.columns[1:]:

                    # Get a dictionary of indices for every breakdown
                    breakdown_dict = \
                        {breakdown.id:
                         {'row': row_index,
                          'column': column_index,
                          'start_row': breakdown.start_row,
                          'start_column': breakdown.start_column,
                          'sheet': breakdown.sheet}
                         for breakdown in breakdown_list}

                    # Get a looping index
                    i = 0
                    # Get a cell value
                    cell_value = '='

                    # For every Breakdown in the list...
                    for key, value in breakdown_dict.items():
                        # Get the cell's column letter, adjusting from
                        # absolute
                        column_string = \
                            get_column_letter(
                                value['start_column'] + value['column']
                                )
                        # Get the cell's row, adjusting from absolute
                        row_string = str(value['start_row'] + value['row'])
                        # Get the cell string
                        cell_string = \
                            (f"'{value['sheet']}'!"
                             f"${column_string}"
                             f"${row_string}")

                        # If the index is 0...
                        if i == 0:
                            # Add the cell_string to the cell_value
                            cell_value += cell_string
                        # Otherwise...
                        else:
                            # Add the cell_string and a '+'
                            cell_value += '+' + cell_string
                        # Iterate the looping index
                        i += 1

                    # Set the cell value
                    self._data.at[index, column] = cell_value

                    # Get the new column index
                    column_index += 1

                # Get the new row index
                row_index += 1

        # Otherwise, raise error
        else:
            raise ValueError(
                'One or more passed breakdowns are of uneven dimensions.'
                )

        # Set the data length
        self.length = len(self._data)

        # Set the breakdown cache
        self._breakdown_cache = {'function': self.merge_breakdowns,
                                 'arguments': {'breakdown_list':
                                               breakdown_list}}

        return None

    # Method to update column references
    @error_logging
    def _update_reference(self):
        """
        Method that updates the Breakdown's reference dictionary.

        Returns
        -------
        None

        """

        # For every column in columns...
        for column in self.columns:

            # Get the start cell's column letter, adjusting from absolute
            col_letter = \
                self._get_column_letter_wrt_start_cell(column,
                                                       self._data,
                                                       self.start_column + 1)

            # Get a Boolean indicating whether the Table has a header
            has_header = False if self.header == '' else True

            # If there is a header...
            if has_header:

                # Get a start row, adjusting from absolute
                start_row = self.start_row + 3

                # Get an end row, adjusting from absolute
                end_row = self.start_row + 2 + self.length

                # Get a range reference, adjusting from absolute
                column_range = (f"'{self._sheet}'!"
                                f"${col_letter}${start_row}:"
                                f"${col_letter}${end_row}")

                # Get a plain range reference
                plain_range = (f"${col_letter}${start_row}:"
                               f"${col_letter}${end_row}")

            # If there isn't a header...
            else:
                # Get a start row, adjusting from absolute
                start_row = self.start_row + 2

                # Get an end row, adjusting from absolute
                end_row = self.start_row + 1 + self.length

                # Get a range reference, adjusting from absolute
                column_range = (f"'{self._sheet}'!"
                                f"${col_letter}${start_row}:"
                                f"${col_letter}${end_row}")

                # Get a plain range reference
                plain_range = column_range.split('!')[-1]

            # Update the reference object
            self._reference[column] = \
                {'column_letter': col_letter,
                 'start_row': start_row,
                 'end_row': end_row,
                 'sheet': self._sheet,
                 'length': self.length,
                 'range': column_range,
                 'plain_range': plain_range}

        return None

    # Method for updating otherwise static Breakdown attributes
    @error_logging
    def _update_breakdown(self):
        """
        Method that updates the current Breakdown.

        Returns
        -------
        None

        """

        # Update the length header
        self.length: int = \
            len(self._data)

        # Update the column header
        self.columns: list[str] = \
            self._data.columns.tolist()

        # Initialize the reference object
        self._reference = {}

        # Try to update the reference
        # NOTE: will not work if there is no valid sheet or start_cell
        try:
            self._update_reference()

        except Exception:
            # logger.info('Failed to update reference.')
            pass

        return None

    """ STATIC METHODS """
    # Static method to get the letter coordinate of a column in a DataFrame
    # with respect to a starting column index
    @staticmethod
    def _get_column_letter_wrt_start_cell(column_name: str,
                                          df: pd.DataFrame,
                                          start_col_index: int) -> str:
        """
        Static method that returns the column letter of a column with respect
        to that column's location within a DataFrame.

        Parameters
        ----------
        column_name : str
            The name of the column to get the index of.
        df : pd.DataFrame
            The DataFrame containing the column of interest.
        start_col_index : int
            The index within the Excel report where the top left of the
            DataFrame will be located.

        Returns
        -------
        column_letter: str
            The letter of the column of interest.
        """

        # Get the current column's letter
        column_letter = \
            get_column_letter(start_col_index +
                              df.columns.get_loc(column_name))

        return column_letter

    # Base conditional aggregate method
    @staticmethod
    def _wrap_conditional_aggregate(inner: str,
                                    outer: str) -> str:
        """
        Static method that returns a basic formula string for Excel using
        passed inner and outer parts.

        Parameters
        ----------
        inner : str
            Component of the formula string defining the Excel formula to
            use (e.g., 'SUM' for '=SUM($A$1:$A$10)')
        outer : str
            Component of the formula string to include within a formula's
            arguments (e.g., '$A$1:$A$10' for '=SUM($A$1:$A$10)').

        Returns
        -------
        formula_string: str
            Complete formula string.
        """

        # If the inner starts with an equals sign...
        if inner[0] == '=':
            # Remove the equals sign
            inner = inner[1:]

        # Otherwise, pass
        else:
            pass

        # Get the formula string
        formula_string = f"={outer}({inner})"

        return formula_string
