#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
COPYRIGHT STATEMENT:

ChromaQuant – A quantification software for complex gas chromatographic data

Copyright (c) 2026, by Julia Hancock
              Affiliation: Dr. Julie Elaine Rorrer
              URL: https://www.rorrerlab.com/

License: BSD 3-Clause License

---

CLASS DEFINITION FOR TABLES

Julia Hancock
Started 01-12-2026

"""

from collections.abc import Callable
import logging
import pandas as pd
from typing import Any, TYPE_CHECKING
if TYPE_CHECKING:
    from ..results import Results
from openpyxl.utils.cell import get_column_letter
from .dataset import DataSet
from ..formula import Formula
from ..logging_and_handling import setup_logger, setup_error_logging
from ..match import match, MatchConfig
from ..utils import get_molecular_weight, get_number_element_atoms

""" LOGGING AND HANDLING """

# Create a logger
logger = logging.getLogger(__name__)

# Format the logger
logger = setup_logger(logger)

# Get an error logging decorator
error_logging = setup_error_logging(logger)

""" CLASS """


# Define the Table class
class Table(DataSet):

    def __init__(self,
                 data_frame: pd.DataFrame | None = None,
                 start_cell: str = '',
                 sheet: str = '',
                 results: Results = None):

        # Create a default DataFrame
        data_frame = data_frame if data_frame is not None else pd.DataFrame()

        # Run DataSet initialization
        super().__init__(data=data_frame,
                         start_cell=start_cell,
                         sheet=sheet,
                         type='Table')

        # Update the table
        self.update_table()

    """ PROPERTIES """
    # Define the reference property
    # Only give reference a getter
    @property
    def reference(self):
        self.update_table()
        return self._reference

    # Redefining properties to include update_table
    # Data properties
    # Getter
    @property
    def data(self) -> pd.DataFrame | None:
        return self._data

    # Setter
    @data.setter
    def data(self, value: pd.DataFrame | None):
        self._data = value
        self.update_table()

    # Deleter
    @data.deleter
    def data(self):
        del self._data
        self._data = pd.DataFrame()
        self.update_table()

    # Sheet properties
    # Getter
    @property
    def sheet(self) -> str:
        return self._sheet

    # Setter
    @sheet.setter
    def sheet(self, value: str):
        if value == '':
            raise ValueError('Table sheet cannot be an empty string.')
        self._sheet = value
        self.update_table()
        if self._mediator is not None:
            self._mediator.update_datasets()

    # Deleter
    @sheet.deleter
    def sheet(self):
        del self._sheet
        self.update_table()
        if self._mediator is not None:
            self._mediator.update_datasets()

    # Start cell properties
    # Getter
    @property
    def start_cell(self) -> str:
        return self._start_cell

    # Setter
    @start_cell.setter
    def start_cell(self, value: str):
        # Try...
        try:
            # Get the cell's absolute indices
            self.start_column, self.start_row = self.get_cell_indices(value)
            # Set the starting cell
            self._start_cell = value
        # If an exception occurs...
        except Exception as e:
            raise ValueError(f'Passed start cell is not valid: {e}')
        self.update_table()
        if self._mediator is not None:
            self._mediator.update_datasets()

    # Deleter
    @start_cell.deleter
    def start_cell(self):
        del self._start_cell
        del self.start_row, self.start_column
        self.update_table()
        if self._mediator is not None:
            self._mediator.update_datasets()

    """ METHODS """
    # Method to add a new column to a table
    @error_logging
    def add_table_column(self,
                         column_name: str,
                         column_values: Any = float('nan')):

        # Set every entry in the column to column_values
        # If column_values is an iterable, this will iterate
        # through it, otherwise it will set every entry to
        # the same value
        self._data[column_name] = column_values

        # Update the table
        self.update_table()

        return None

    # Method to add a new column to a table by passing a list
    # returned from a function, useful for Python operations on
    # one or more columns in the DataFrame
    @error_logging
    def add_table_column_from_function(self,
                                       column_name: str,
                                       function: Callable[..., list[Any]],
                                       *args: Any,
                                       **kwargs: Any):

        # Get the output of a function using passed arguments
        function_output = function(*args, **kwargs)

        # Add the function output as a column to the table
        self.add_table_column(column_name, function_output)

        return None

    # Method to add a molecular weight column based
    # on another column with chemical formulas
    @error_logging
    def add_molecular_weight_column(self,
                                    formula_column_name: str,
                                    new_column_name: str = 'Molecular weight'):

        # Get the formula column as a list
        formula_list = self._data[formula_column_name].tolist()

        # Create a new molecular weight column by passing this
        # formula column and the molecular weight function
        self.add_table_column_from_function(new_column_name,
                                            get_molecular_weight,
                                            formula=formula_list)

        return None

    # Method to add an element count column based
    # on another column with chemical formulas
    @error_logging
    def add_element_count_column(self,
                                 formula_column_name: str,
                                 element: str,
                                 new_column_name: str = ''):

        # Get a new column name based on the element if empty
        new_column_name = \
            element if new_column_name == '' else new_column_name

        # Get the formula column as a list
        formula_list = self._data[formula_column_name].tolist()

        # Create a new molecular weight column by passing this
        # formula column and the molecular weight function
        self.add_table_column_from_function(new_column_name,
                                            get_number_element_atoms,
                                            formula=formula_list,
                                            element=element)

        return None

    # Method to add a formula column to the DataFrame
    @error_logging
    def add_formula(self,
                    formula: Formula,
                    ):

        return None

    # Method to get the insert string for a given column
    @error_logging
    def insert(self, column: str, range: bool = False) -> str:

        # If range is True...
        if range is True:
            # Set insert to a Table's column range insert
            insert = f'|table: {self.id}, key: {column}, range: True|'

        # Otherwise...
        else:
            # Set insert to Table's column insert
            insert = f'|table: {self.id}, key: {column}|'

        return insert

    # Method to import data
    @error_logging
    def import_csv_data(self, path: str, **kwargs: Any):
        """Reads .csv at passed path and sets self._data to result.

        Parameters
        ----------
        path : str
            Full path to the data file of interest, including file name.
        data_key : str
            Name to assign imported data when using the "data" object.

        Returns
        -------
        None

        """

        # If the user did not pass an index_col value...
        if 'index_col' not in kwargs:
            # Set index_col to False
            kwargs['index_col'] = False
        # Otherwise, pass
        else:
            pass

        # Read data from the provided directory
        self._data = pd.read_csv(path, **kwargs)

        return None

    # Method to match one dataframe to current data
    def match(self,
              import_DF: pd.DataFrame,
              match_config: MatchConfig) -> pd.DataFrame:

        return match(self._data, import_DF, match_config)

    # Method to update column references
    @error_logging
    def update_reference(self):

        # For every column in columns...
        for column in self.columns:

            # Get the start cell's column letter, adjusting from absolute
            col_letter = \
                self.get_column_letter_wrt_start_cell(column,
                                                      self._data,
                                                      self.start_column + 1)

            # Get a start row, adjusting from absolute
            start_row = self.start_row + 1

            # Get an end row, adjusting from absolute
            end_row = self.start_row + 1 + self.length

            # Get a range reference, adjusting from absolute
            column_range = (f"'{self._sheet}'!"
                            f"${col_letter}${start_row + 1}:"
                            f"${col_letter}${end_row}")

            # Update the reference object
            self._reference[column] = \
                {'column_letter': col_letter,
                 'start_row': start_row,
                 'end_row': end_row,
                 'sheet': self._sheet,
                 'length': self.length,
                 'range': column_range}

        return None

    # Method for updating otherwise static Table attributes
    @error_logging
    def update_table(self):

        # Update the length header
        self.length = \
            len(self._data)

        # Update the column header
        self.columns: list[str] = \
            self._data.columns.tolist()
        # Initialize the reference object
        self._reference = {}

        # Try to update the reference
        # NOTE: will not work if there is no valid sheet or start_cell
        try:
            self.update_reference()

        except Exception:
            logger.info('Failed to update reference.')
            pass

        return None

    """ STATIC METHODS """
    # Static method to get the letter coordinate of a column in a DataFrame
    # with respect to a starting column index
    @staticmethod
    def get_column_letter_wrt_start_cell(column: str,
                                         df: pd.DataFrame,
                                         start_col_index: int) -> str:

        # Get the current column's letter
        col_letter = \
            get_column_letter(start_col_index +
                              df.columns.get_loc(column))

        return col_letter
