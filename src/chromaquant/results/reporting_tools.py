#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

This submodule contains functions used in Excel reporting, particularly
by the Results class (see Results).

"""

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from pandas import ExcelWriter
from ..chart import Chart
from ..data import Breakdown, Table, Value
from ..theme.theme import CellStyle

""" FUNCTIONS """


# Function to format a given cell using a passed CellStyle
def format_cell(cell: Cell,
                group: CellStyle):
    """
    Format a cell in Excel.

    Parameters
    ----------
    cell : Cell
        Cell to format.
    group : CellStyle
        Style to format the cell by.

    Returns
    -------
    None

    """
    # Change all settings
    cell.font = group.font
    cell.fill = group.fill
    cell.border = group.border
    cell.alignment = group.alignment
    cell.protection = group.protection
    cell.number_format = group.number_format

    return None


# Function to format a table or breakdown
def format_multicell_dataset(dataset: Table | Breakdown,
                             workbook: Workbook):
    """
    Format a Table or Breakdown in Excel.

    Parameters
    ----------
    dataset : Table | Breakdown
        Dataset to format.
    workbook : Workbook
        Active openpyxl workbook.

    Returns
    -------
    None

    """

    # Get the dataset's footprint
    footprint = dataset.footprint

    # If dataset's sheet does not exist in workbook...
    if dataset.sheet not in workbook.sheetnames:
        # Create it
        workbook.create_sheet(dataset.sheet)

    # Open the dataset's sheet
    sheet = workbook[dataset.sheet]

    # If the dataset has a header...
    if dataset.header != '':

        # For every cell in the header footprint...
        # NOTE: assumes header footprint covers only one row
        for cell in sheet[footprint['header']][0]:
            # Write the header value to cell
            cell.value = dataset.header

        # Format the header range using the dataset's theme attribute's header
        # style
        format_range(sheet, footprint['header'], dataset.theme.header)

        # Get the number of columns in the DataSet's data
        num_cols = dataset.data.shape[1]

        # Merge the header cell with adjacent cells
        # to center the header across the DataFrame
        sheet.merge_cells(start_row=dataset.start_row + 1,
                          start_column=dataset.start_column + 1,
                          end_row=dataset.start_row + 1,
                          end_column=dataset.start_column + num_cols)

    # Otherwise, pass
    else:
        pass

    # Format the subheader range using the subheader style
    format_range(sheet, footprint['subheader'], dataset.theme.subheader)

    # For every column in the body footprint...
    for column, range in footprint['body'].items():
        # Format the column range using the subheader body
        format_range(sheet, range, dataset.theme.body)

    return None


# Function to format a range of cells
def format_range(sheet: Worksheet,
                 cell_range: str,
                 group: CellStyle):

    # For every row in the range...
    for row in sheet[cell_range]:

        # For every cell in the row...
        for cell in row:

            # Format the cell
            format_cell(cell, group)

    return None


# Function to write a Breakdown to Excel
def report_breakdown(breakdown: Breakdown,
                     writer: ExcelWriter):
    """
    Writes a Pandas Breakdown to Excel using passed writer.

    Parameters
    ----------
    breakdown : Breakdown
        Pandas Breakdown to export.
    writer : ExcelWriter
        Pandas ExcelWriter used in exporting.

    Returns
    ----------
    None

    """
    # Get the start row based on whether there is a header or not
    start_row = \
        breakdown.start_row if breakdown.header == '' \
        else breakdown.start_row + 1

    # Write the passed DataFrame to passed path
    breakdown.data.to_excel(writer,
                            sheet_name=breakdown.sheet,
                            startcol=breakdown.start_column,
                            startrow=start_row,
                            index=False)

    return None


# Function to report a Chart to Excel
def report_chart(chart: Chart,
                 workbook: Workbook):
    """
    Writes a Chart to Excel.

    Parameters
    ----------
    chart : Chart
        Chart to report.
    workbook : Workbook
        Active openpyxl workbook.

    Returns
    -------
    None

    """

    # If Chart's sheet does not exist in workbook...
    if chart.sheet not in workbook.sheetnames:
        # Create it
        workbook.create_sheet(chart.sheet)

    # Get the anchor cell
    anchor = \
        get_column_letter(chart.start_column + 1) + str(chart.start_row + 1)

    # Get the Chart's worksheet
    ws = workbook[chart.sheet]

    # Add the Chart to its worksheet
    ws.add_chart(chart.base, anchor)

    return None


# Function to write a Table to Excel
def report_table(table: Table,
                 writer: ExcelWriter):
    """
    Writes a Table to Excel.

    Parameters
    ----------
    table: Table
        Table to export.
    writer : ExcelWriter
        Pandas ExcelWriter used in exporting.

    Returns
    -------
    None

    """

    # Get the start row based on whether there is a header or not
    start_row = \
        table.start_row if table.header == '' \
        else table.start_row + 1

    # Write the passed DataFrame to passed path
    table.data.to_excel(writer,
                        sheet_name=table.sheet,
                        startcol=table.start_column,
                        startrow=start_row,
                        index=False)

    return None


# Function to write a Value to Excel
def report_value(value: Value,
                 workbook: Workbook):
    """
    Writes a Value to Excel.

    Parameters
    ----------
    value: Value
        Value to export.
    workbook : Workbook
        Active openpyxl workbook.

    Returns
    -------
    None

    """

    # If Value's sheet does not exist in workbook...
    if value.sheet not in workbook.sheetnames:
        # Create it
        workbook.create_sheet(value.sheet)

    # Open the Value's sheet
    sheet = workbook[value.sheet]

    # Get the start cell
    start_cell = \
        get_column_letter(value.start_column + 1) + str(value.start_row + 1)

    # Get the cell after
    second_cell = \
        get_column_letter(value.start_column + 1) + str(value.start_row + 2)

    # If there is a header...
    if value.header != '':
        # Write the header to the start cell
        sheet[start_cell] = value.header
        # Format the cell using the value's theme's header style
        format_cell(sheet[start_cell], value.theme.header)
        # Write the value to the second cell
        sheet[second_cell] = value.data
        # Format the cell using the value's theme's body style
        format_cell(sheet[second_cell], value.theme.body)

    # Otherwise...
    else:
        # Write the value to the start cell
        sheet[start_cell] = value.data
        # Format the cell using the value's theme's body style
        format_cell(sheet[start_cell], value.theme.body)

    return None


# Function to set all columns to a default width
def set_default_col_widths(workbook: Workbook):
    """
    Sets all columns in an active openpyxl workbook to a default width.

    Parameters
    ----------
    workbook : Workbook
        Active openpyxl workbook.

    Returns
    -------
    None

    """

    # For every worksheet in the workbook...
    for sheet in workbook.worksheets:
        # For every column in the sheet...
        for column in sheet.iter_cols():
            # Get the maximum length of all values in the column
            column_width = max(len(str(cell.value)) for cell in column)
            # If the column_width is above 25, set to 25
            column_width = 25 if column_width > 25 else column_width
            # If the column_width is less than 8, set to 8
            column_width = 8 if column_width < 8 else column_width
            # Get the column's index
            col_index = column[0].column
            # Set the column's width to the minimum required to fit all values
            sheet.column_dimensions[get_column_letter(col_index)].width = \
                column_width

    return None
