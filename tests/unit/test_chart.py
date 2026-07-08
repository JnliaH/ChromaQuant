#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

UNIT TESTING FOR CHART

"""

import chromaquant as cq
from openpyxl import Workbook, load_workbook
from openpyxl.chart import ScatterChart, BarChart

""" TEST CLASS """


class TestChart:

    # Test Chart creation with openpyxl
    def test_chart_init(self):

        # Create a new workbook
        wb = Workbook()

        # Get the default sheet
        ws = wb['Sheet']

        # Create a new Chart
        new_chart = \
            cq.Chart(
                ScatterChart(),
                )

        # Add a theme to the Chart
        new_chart.theme = cq.Theme()

        # Add the chart to the worksheet
        ws.add_chart(new_chart.base, 'C5')

        # Save the workbook
        wb.save('./tests/unit/test_chart.xlsx')

        # Read the workbook
        wb = load_workbook('./tests/unit/test_chart.xlsx')

        # Get the default sheet
        ws = wb['Sheet']

        # Assert that the workbook contains at least one chart
        assert ws._charts

    # Test Chart creation using Table data
    def test_table_chart(self):

        # Create a new Table
        table = cq.Table(start_cell='B2',
                         header='Table')

        # Add data to Table
        table.data['x'] = [1, 2, 3, 4, 5]
        table.data['A'] = [10, 20, 35, 28, 46]
        table.data['B'] = [8, 2, 14, 26, 52]
        table.data['C'] = [80, 72, 64, 31, 25]

        # Add Table to Results
        results = cq.Results()
        results.add_table(table)

        # Create a new Chart
        chart = cq.Chart(chart=BarChart(),
                         theme=cq.Theme(),
                         anchor='$C$15')

        # Add references to data
        chart.indep_column = table.column_id('x')
        chart.data_columns = [
            table.column_id('A'),
            table.column_id('B'),
            table.column_id('C')
        ]

        # Add the Chart to Results
        results.add_chart(chart)

        # Report the results
        results.report_results('./tests/unit/test_table_chart.xlsx')
